"""Build the canonical segment dataset from the monthly platform exports.

Reads the two source `.xlsx` exports and normalises both platforms into one
canonical `data/segments.csv` that every downstream consumer (brief pipeline,
Prism Assistant, deck builder) reads. Runtime code only ever touches the clean
CSV — no `.xlsx` parsing happens in the request path. See PRD #96 / slice #97.

Sources:
  * Permutive behavioural segments — sheet "Permutive Segments" in
    `~/Documents/AP Audiences.xlsx`. Category lives in the `Name` prefix
    ("Food & Drink - …"); engagement is encoded as a `(R/O/F & window)` suffix
    where R = browsed >=2x, F = browsed >=5x, O = browsed >=once.
  * AP first-party / DFP intent segments — sheet "AP Cleaned Segment Names" in
    `~/Downloads/AP Audience List.xlsx` (the latest export, which carries the
    `Size` reach column and a clean `Segment Category` column).

Normalisation applied at build time:
  * Unified schema (see SEGMENT_FIELDS).
  * Permutive category parsed from the Name prefix; AP category from its column.
  * R-variant collapse: group Permutive rows by `Clean Audience Name`, keep the
    R (Regular, >=2x) variant, falling back R -> F -> O.
  * Exclude any segment with reach <= 0 or blank (un-pitchable).
  * Derive a cleaned plain-English descriptor from the raw description.

Usage:
    python3 scripts/build_segments.py
    python3 scripts/build_segments.py --permutive PATH --ap PATH --out PATH

Re-run whenever a fresh monthly export lands, then commit the updated CSV.
"""
import argparse
import csv
import os
import re

import openpyxl

HERE = os.path.dirname(__file__)
DEFAULT_PERMUTIVE = os.path.expanduser("~/Documents/AP Audiences.xlsx")
DEFAULT_AP = os.path.expanduser("~/Downloads/AP Audience List.xlsx")
DEFAULT_OUT = os.path.normpath(os.path.join(HERE, "..", "data", "segments.csv"))

PERMUTIVE_SHEET = "Permutive Segments"
AP_SHEET = "AP Cleaned Segment Names"

# Canonical output schema. Order is the CSV column order.
SEGMENT_FIELDS = [
    "segment_name",
    "platform",
    "reach",
    "category",
    "description",
    "plain_english",
    "code",
    "frequency",
    "window",
    "icon_key",
]

# Engagement letter -> (rank for R->F->O collapse, human descriptor verb).
# R beats F beats O; a row with no recognised letter ranks lowest.
_VARIANT_RANK = {"R": 3, "F": 2, "O": 1, "": 0}
_VARIANT_VERB = {
    "R": "Regularly browses",
    "F": "Frequently browses",
    "O": "Occasionally browses",
    "": "Browses",
}

# Matches the `(R & 90 Days)` engagement suffix anywhere in the raw Name.
# Only R/O/F are treated as real variants; everything else is noise.
_VARIANT_RE = re.compile(r"\(\s*([ROF])\s*&\s*([^)]*?)\s*\)")
_TOPIC_RE = re.compile(r"browsed for (.*?)\s+content", re.IGNORECASE)
_AP_PREFIX_RE = re.compile(r"^we define these users as those who\s+", re.IGNORECASE)

# Category strings that differ only by casing/spelling across the two exports.
# Keys are lower-cased; values are the canonical display form.
_CATEGORY_CANON = {
    "arts & entertainment": "Arts & Entertainment",
    "food & drink": "Food & Drink",
    "sport": "Sports",
    "sports": "Sports",
}

# Window strings that differ only by casing/spelling.
_WINDOW_CANON = {
    "all time": "All Time",
    "all times": "All Time",
    "90 days": "90 Days",
    "90days": "90 Days",
    "30 days": "30 Days",
    "7 days": "7 Days",
    "1 month": "1 Month",
    "6 months": "6 Months",
    "current session": "Current Session",
}


def _clean(value):
    """Collapse whitespace and strip; return '' for None."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def to_int_reach(size):
    """Return reach as a positive int, or None if blank/zero/negative/garbage."""
    if size is None or size == "":
        return None
    try:
        value = int(float(size))
    except (TypeError, ValueError):
        return None
    return value if value > 0 else None


def canonical_category(raw):
    """Normalise a category string, folding known case/spelling variants."""
    cleaned = _clean(raw)
    if not cleaned:
        return "Uncategorised"
    return _CATEGORY_CANON.get(cleaned.lower(), cleaned)


def icon_key_for(category):
    """Deterministic slug used as the icon lookup key (refined in slice #100)."""
    slug = re.sub(r"[^a-z0-9]+", "-", canonical_category(category).lower())
    return slug.strip("-") or "uncategorised"


def parse_permutive_category(name):
    """Category is the prefix before the first ' - ' in the raw Name."""
    cleaned = _clean(name)
    if " - " in cleaned:
        return canonical_category(cleaned.split(" - ", 1)[0])
    return "Uncategorised"


def extract_variant(name):
    """Return (letter, window) from a `(R/O/F & window)` suffix, else ('', '')."""
    match = _VARIANT_RE.search(name or "")
    if not match:
        return "", ""
    return match.group(1), normalise_window(match.group(2))


def normalise_window(window):
    """Fold known casing/spelling variants of the engagement window."""
    cleaned = _clean(window)
    if not cleaned:
        return ""
    return _WINDOW_CANON.get(cleaned.lower(), cleaned)


def permutive_plain_english(description, letter):
    """Derive a clean descriptor from the raw 'browsed for X content' text."""
    topic_match = _TOPIC_RE.search(description or "")
    verb = _VARIANT_VERB.get(letter, "Browses")
    if topic_match:
        topic = _clean(topic_match.group(1))
        return "%s %s content" % (verb, topic)
    return verb + " related content"


def ap_plain_english(description):
    """Strip the boilerplate 'We define these users as those who' lead-in."""
    cleaned = _clean(description)
    stripped = _AP_PREFIX_RE.sub("", cleaned)
    if not stripped:
        return cleaned
    return stripped[0].upper() + stripped[1:]


def normalise_permutive_row(row):
    """Map a raw Permutive tuple to a canonical dict, or None if un-pitchable.

    Row order: Name, Clean Audience Name, Code, Description, Size.
    Returns a dict with an extra private `_clean_name`/`_rank` for collapse.
    """
    name, clean_name, code, description, size = (list(row) + [None] * 5)[:5]
    reach = to_int_reach(size)
    if reach is None:
        return None
    letter, window = extract_variant(name)
    category = parse_permutive_category(name)
    segment_name = _clean(clean_name) or _clean(name)
    return {
        "segment_name": segment_name,
        "platform": "Permutive",
        "reach": reach,
        "category": category,
        "description": _clean(description),
        "plain_english": permutive_plain_english(description, letter),
        "code": _clean(int(code)) if isinstance(code, float) else _clean(code),
        "frequency": letter,
        "window": window,
        "icon_key": icon_key_for(category),
        "_clean_name": segment_name.lower(),
        "_rank": _VARIANT_RANK.get(letter, 0),
    }


def normalise_ap_row(row, headers):
    """Map a raw AP row (with headers) to a canonical dict, or None if blank."""
    record = dict(zip(headers, row))
    reach = to_int_reach(record.get("Size"))
    if reach is None:
        return None
    category = canonical_category(record.get("Segment Category"))
    description = _clean(record.get("Cleaned Segment Description"))
    return {
        "segment_name": _clean(record.get("Cleaned name")) or _clean(record.get("Name")),
        "platform": "AP",
        "reach": reach,
        "category": category,
        "description": description,
        "plain_english": ap_plain_english(description),
        "code": _clean(record.get("Cleaned Segment Code")),
        "frequency": "",
        "window": "",
        "icon_key": icon_key_for(category),
    }


def collapse_permutive_variants(rows):
    """Keep one row per clean name, favouring R -> F -> O, then higher reach."""
    best = {}
    for row in rows:
        key = row["_clean_name"]
        current = best.get(key)
        if current is None or (row["_rank"], row["reach"]) > (current["_rank"], current["reach"]):
            best[key] = row
    collapsed = []
    for row in best.values():
        row.pop("_clean_name", None)
        row.pop("_rank", None)
        collapsed.append(row)
    return collapsed


def load_permutive(path):
    """Read + normalise + collapse the Permutive sheet into canonical rows."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[PERMUTIVE_SHEET]
        raw = [normalise_permutive_row(r) for r in sheet.iter_rows(min_row=2, values_only=True)]
    finally:
        workbook.close()
    return collapse_permutive_variants([r for r in raw if r is not None])


def load_ap(path):
    """Read + normalise the AP sheet into canonical rows."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[AP_SHEET]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    headers = [_clean(h) for h in rows[0]]
    out = [normalise_ap_row(r, headers) for r in rows[1:]]
    return [r for r in out if r is not None]


def write_csv(rows, out_path):
    with open(out_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=SEGMENT_FIELDS)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in SEGMENT_FIELDS})


def main():
    parser = argparse.ArgumentParser(description="Build the canonical segments.csv.")
    parser.add_argument("--permutive", default=DEFAULT_PERMUTIVE, help="Permutive .xlsx export")
    parser.add_argument("--ap", default=DEFAULT_AP, help="AP .xlsx export (with Size column)")
    parser.add_argument("--out", default=DEFAULT_OUT, help="Output CSV path")
    args = parser.parse_args()

    permutive = load_permutive(args.permutive)
    ap = load_ap(args.ap)
    rows = permutive + ap
    write_csv(rows, args.out)

    print("Wrote %d segments to %s" % (len(rows), args.out))
    print("  Permutive: %d (after R->F->O collapse)" % len(permutive))
    print("  AP:        %d" % len(ap))
    print("  Note: reach is per-segment and MUST NOT be summed (segments overlap).")


if __name__ == "__main__":
    main()
